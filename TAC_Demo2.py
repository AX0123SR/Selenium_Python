import time
import openpyxl
from datetime import date
from selenium import webdriver
import win32com.client
from selenium.webdriver.support.ui import Select


class HappyHour:
    driver = webdriver.Chrome(executable_path="C:/Users/ayussriv/Downloads/chromedriver_win32/chromedriver.exe")




    def login(self):
        #self.driver = webdriver.Chrome(executable_path="C://Users//ayussriv//Downloads//chromedriver//chromedriver.exe")

        self.driver.get("https://tac.rdisoftware.com/Account/Login")
        time.sleep(2)
        self.driver.maximize_window()
        time.sleep(2)
        shell = win32com.client.Dispatch("WScript.Shell")  # win32com provides access to automation modules.

        ''' First Login credentials for TAC '''
        shell.Sendkeys("ayush.a.srivastava")  # SendKeys are used to send input to authentication alert boxes
        time.sleep(2)  # pauses execution for 2 seconds
        shell.Sendkeys("{TAB}")
        time.sleep(2)
        shell.Sendkeys("S41&14s_")
        time.sleep(2)
        shell.Sendkeys("{ENTER}")
        time.sleep(2)

       # self.driver.find_element_by_xpath("//button[text()='OK']").click()
        #time.sleep(1)

        ''' Second Login credentials for TAC '''
        self.driver.find_element_by_id("userNameTBox").send_keys("asrivastava")
        time.sleep(1)
        self.driver.find_element_by_id("psswdTBox").send_keys("@/Ayush123")
        time.sleep(1)
        self.driver.find_element_by_id("loginPerform").click()
        time.sleep(2)

       # self.driver.find_element_by_xpath("//button[text()='OK']").click()
        #time.sleep(1)

        ''' EMA '''
        self.driver.find_element_by_id("EMAMenuID").click()
        time.sleep(1)
        ''' Test Suite '''
        self.driver.find_element_by_id("TSMenuID").click()
        ''' Markets '''
        self.driver.find_element_by_id("select2-MarketId-container").click()
        time.sleep(1)
        ''' Select US '''
        self.driver.find_element_by_xpath("//li[text()='US']").click()
        time.sleep(1)
        ''' Environments '''
        self.driver.find_element_by_id("select2-EnvironmentId-container").click()
        time.sleep(1)
        ''' Select NpSharp '''
        self.driver.find_element_by_xpath("//li[text()='NpSharp']").click()
        time.sleep(1)
        ''' Module name '''
        self.driver.find_element_by_id("ModuleName").click()
        time.sleep(1)
        self.driver.find_element_by_id("ModuleName").send_keys("Happy Hour Drink")
        time.sleep(1)
        ''' CLick on Filter button '''
        self.driver.find_element_by_id("filterButton").click()
        time.sleep(1)

    def HappyHour_Set_1(self):
        ''' Open url in new window '''
        self.driver.execute_script("window.open('');")
        self.driver.switch_to.window(self.driver.window_handles[1])
        ''' Happy hour Set 1 '''
        self.driver.get("https://tac.rdisoftware.com/TestSuite/Index/18545?moduleId=1929")
        time.sleep(3)

    def HappyHour_Set_2(self):
        ''' Open url in new window '''
        self.driver.execute_script("window.open('');")
        self.driver.switch_to.window(self.driver.window_handles[2])
        ''' Happy hour Set 2 '''
        self.driver.get("https://tac.rdisoftware.com/TestSuite/Index/18546?moduleId=1929")
        time.sleep(3)

    def HappyHour_Set_3(self):
        ''' Open url in new window '''
        self.driver.execute_script("window.open('');")
        self.driver.switch_to.window(self.driver.window_handles[3])
        ''' Happy hour Set 3 '''
        self.driver.get("https://tac.rdisoftware.com/TestSuite/Index/18547?moduleId=1929")
        time.sleep(3)

    def HappyHour_Set_4(self):
        ''' Open url in new window '''
        self.driver.execute_script("window.open('');")
        self.driver.switch_to.window(self.driver.window_handles[4])
        ''' Happy hour Set 4 '''
        self.driver.get("https://tac.rdisoftware.com/TestSuite/Index/18548?moduleId=1929")
        time.sleep(3)

    def HappyHour_Set_12(self):
        ''' Open url in new window '''
        self.driver.execute_script("window.open('');")
        self.driver.switch_to.window(self.driver.window_handles[5])
        ''' Happy hour Set 12 '''
        self.driver.get("https://tac.rdisoftware.com/TestSuite/Index/18556?moduleId=1929")
        time.sleep(3)



    def ScheduleTestSuite(self):
        #self.driver = webdriver.Chrome(executable_path="C://Users//ayussriv//Downloads//chromedriver//chromedriver.exe")

        ''' Schedule Test Suite'''
        self.driver.find_element_by_id("btnScheduleTestSuite").click()
        time.sleep(2)

        ''' Scroll down till the end '''
        self.driver.execute_script("window.scrollBy(0,document.body.scrollHeight)")

        ''' Development Phase '''
        drop = Select(self.driver.find_element_by_id("selSchedulePhases"))
        time.sleep(1)
        drop.select_by_value("1")
        ''' 6.1.33 Major '''
        drop = Select(self.driver.find_element_by_id("selScheduleMajors"))
        time.sleep(1)
        drop.select_by_value("24")
        ''' MR01 Minor '''
        drop = Select(self.driver.find_element_by_id("selScheduleMinors"))
        time.sleep(1)
        drop.select_by_value("1")
        ''' Quality QR01 '''
        drop = Select(self.driver.find_element_by_id("selScheduleQuality"))
        time.sleep(1)
        drop.select_by_value("1")
        time.sleep(3)

        ''' DatePicker '''
        datepicker = self.driver.find_element_by_id("scheduleDatePicker")
        datepicker.click()
        time.sleep(1)
        today = date.today()
        # dd/mm/YY
        d1 = int(today.strftime("%d"))
        print("d1 =", d1)

        a = d1 + 1
        print(a)
        if a == 1:
            self.driver.find_element_by_xpath("//div[text()='1']").click()
        if a == 2:
            self.driver.find_element_by_xpath("//div[text()='2']").click()
        if a == 3:
            self.driver.find_element_by_xpath("//div[text()='3']").click()
        if a == 4:
            self.driver.find_element_by_xpath("//div[text()='4']").click()
        if a == 5:
            self.driver.find_element_by_xpath("//div[text()='5']").click()
        if a == 6:
            self.driver.find_element_by_xpath("//div[text()='6']").click()
        if a == 7:
            self.driver.find_element_by_xpath("//div[text()='7']").click()
        if a == 8:
            self.driver.find_element_by_xpath("//div[text()='8']").click()
        if a == 9:
            self.driver.find_element_by_xpath("//div[text()='9']").click()
        if a == 10:
            self.driver.find_element_by_xpath("//div[text()='10']").click()
        if a == 11:
            self.driver.find_element_by_xpath("//div[text()='11']").click()
        if a == 12:
            self.driver.find_element_by_xpath("//div[text()='12']").click()
        if a == 13:
            self.driver.find_element_by_xpath("//div[text()='13']").click()
        if a == 14:
            self.driver.find_element_by_xpath("//div[text()='14']").click()
        if a == 15:
            self.driver.find_element_by_xpath("//div[text()='15']").click()
        if a == 16:
            self.driver.find_element_by_xpath("//div[text()='16']").click()
        if a == 17:
            self.driver.find_element_by_xpath("//div[text()='17']").click()
        if a == 18:
            self.driver.find_element_by_xpath("//div[text()='18']").click()
        if a == 19:
            self.driver.find_element_by_xpath("//div[text()='19']").click()
        if a == 20:
            self.driver.find_element_by_xpath("//div[text()='20']").click()
        if a == 21:
            self.driver.find_element_by_xpath("//div[text()='21']").click()
        if a == 22:
            self.driver.find_element_by_xpath("//div[text()='22']").click()
        if a == 23:
            self.driver.find_element_by_xpath("//div[text()='23']").click()
        if a == 24:
            self.driver.find_element_by_xpath("//div[text()='24']").click()
        if a == 25:
            self.driver.find_element_by_xpath("//div[text()='25']").click()
        if a == 26:
            self.driver.find_element_by_xpath("//div[text()='26']").click()
        if a == 27:
            self.driver.find_element_by_xpath("//div[text()='27']").click()
        if a == 28:
            self.driver.find_element_by_xpath("//div[text()='28']").click()
        if a == 29:
            self.driver.find_element_by_xpath("//div[text()='29']").click()
        if a == 30:
            self.driver.find_element_by_xpath("//div[text()='30']").click()
        if a == 31:
            self.driver.find_element_by_xpath("//div[text()='31']").click()

        s = datepicker.get_attribute('value')
        print("Date entered is: ")
        print(s)

        path = "C:/Users/ayussriv/Documents/Demo1.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active  # or workbook.get_sheet_by_name("Sheet 1")

        a1 = sheet["A1"]
        self.driver.find_element_by_id("buildNameInputId").send_keys(a1.value)
        a2 = sheet["A2"]
        self.driver.find_element_by_id("buildPathInputId").send_keys(a2.value)
        #a3 = sheet["A3"]
        #self.driver.find_element_by_id("inputScheduleLocalMachineIP").send_keys(a3.value)



    def VM1(self):
        path = "C:/Users/ayussriv/Documents/Demo1.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active  # or workbook.get_sheet_by_name("Sheet 1")
        d4 = sheet["D4"]
        self.driver.find_element_by_id("inputScheduleLocalMachineIP").send_keys(d4.value)

    def VM2(self):
        path = "C:/Users/ayussriv/Documents/Demo1.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active  # or workbook.get_sheet_by_name("Sheet 1")
        d5 = sheet["D5"]
        self.driver.find_element_by_id("inputScheduleLocalMachineIP").send_keys(d5.value)

    def VM3(self):
        path = "C:/Users/ayussriv/Documents/Demo1.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active  # or workbook.get_sheet_by_name("Sheet 1")
        d6 = sheet["D6"]
        self.driver.find_element_by_id("inputScheduleLocalMachineIP").send_keys(d6.value)

    def VM4(self):
        path = "C:/Users/ayussriv/Documents/Demo1.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active  # or workbook.get_sheet_by_name("Sheet 1")
        d7 = sheet["D7"]
        self.driver.find_element_by_id("inputScheduleLocalMachineIP").send_keys(d7.value)

    def VM5(self):
        path = "C:/Users/ayussriv/Documents/Demo1.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active  # or workbook.get_sheet_by_name("Sheet 1")
        d8 = sheet["D8"]
        self.driver.find_element_by_id("inputScheduleLocalMachineIP").send_keys(d8.value)

    def Schedule(self):
        self.driver.find_element_by_id("scheduleButton").click()



h = HappyHour()
h.login()

h.HappyHour_Set_1()
h.ScheduleTestSuite()
h.VM1()
h.Schedule()

h.HappyHour_Set_2()
h.ScheduleTestSuite()
h.VM2()
h.Schedule()

h.HappyHour_Set_3()
h.ScheduleTestSuite()
h.VM3()
h.Schedule()

h.HappyHour_Set_4()
h.ScheduleTestSuite()
h.VM4()
h.Schedule()

h.HappyHour_Set_12()
h.ScheduleTestSuite()
h.VM5()
h.Schedule()




