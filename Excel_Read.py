'''
Read data from Excel using Openpyxl
1) Install Openpyxl
2) import Openpyxl
3) Give the path for excel file
4) load the workbook
5) open the active excel sheet
6) find no of rows and columns
7) print values of rows and columns
'''

import openpyxl
path = "C:/Users/ayussriv/Documents/Demo.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active  # or workbook.get_sheet_by_name("Sheet 1")

row = sheet.max_row
col = sheet.max_column
print(row,col)

for r in range(1,row+1):
    for c in range(1,col+1):
        print(sheet.cell(row=r,column=c).value, end="    ")
    print()