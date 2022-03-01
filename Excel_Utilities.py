'''
Creating a file for defining modules for rows,columns,read and write into excel file
'''

import openpyxl

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)

def ReadData(file,sheetName,row_no,column_no):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=row_no, column=column_no).value

def WriteData(file,sheetName,row_no,column_no,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=row_no, column=column_no).value = data
    workbook.save(file)

